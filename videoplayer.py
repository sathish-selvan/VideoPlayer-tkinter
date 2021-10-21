from tkinter import *
import  tkinter.filedialog as filedialog
from tkinter import messagebox
import pyautogui
import os
from PIL import Image
import openpyxl
import vlc
import time
# from moviepy.editor import VideoFileClip
import glob
from openpyxl.styles import Alignment


# To get the  window size
w,h = pyautogui.size()
root = Tk()
root.geometry(f"{w}x{h-100}")
## defects code form the code.txt file
global defects
defects = []

# geting defects codes from the text file
with open("codes.txt",'r') as codes:
    code = codes.readlines()
    for i in code:
        try:
            a = i.replace("\n","")
        except:
            pass
        defects.append(a)
defects.sort()
# inserting codes in the list box

global player, you_cannot_pass
path_to_defects_folder = ""
path_to_video = ""
you_cannot_pass = False

# Vlc instance
Instance = vlc.Instance()
player = Instance.media_player_new()
# player.set_playback_mode(vlc.PlaybackMode.loop)

####### ALL THE FUNCTIONS HERE #####
#Choosing the defect path
def add_path_folder():
    global path_to_defects_folder
    path_to_defects_folder = filedialog.askdirectory(initialdir =os.getcwd(), title ="Choose a Defects folder")
    
#Choosing video path
def add_video_file():
    global path_to_video
    path_to_video = filedialog.askopenfilename(initialdir =os.getcwd(), title ="Choose a video file")

#Search bar function
def var_changed():
    val = search.get()
    val = "".join(val.split())
    

    if val == "":
        mylist.delete(0,END)
        for i in defects:
            mylist.insert(END, i) 
    else:
        mylist.delete(0,END)
        val = val.upper()
        for i in defects:
            if val[:] == i[:len(val)]:
                mylist.insert(END, i) 
    _var.set(val)

#Convertion of png to jpg
def jpg_convertion(path):
    images = glob.glob1(path, "*.png")
    for i in images:
        img_png = Image.open(os.path.join(path,i))
        img_png.save(os.path.join(path,i.split(".")[0]+".jpg"))
        os.remove(os.path.join(path,i))


#Screen_shot function
def selected_item(*args):
    global path_to_defects_folder
    code_name = ""
    if path_to_defects_folder == "":
        messagebox.showwarning("Warning","Please select the Defects Folder Directory")
    else:
        for i in mylist.curselection():
            code_name = mylist.get(i)
        if code_name == "":
            messagebox.showwarning("Warning","Please select the Defect CODE")
        else:
            contents = os.listdir(path_to_defects_folder)
            if code_name in contents:
                path_to_images = os.path.join(path_to_defects_folder, code_name)
            else:
                path_to_images = os.path.join(path_to_defects_folder, code_name)
                os.mkdir(path_to_images)
            global player
            
            player.video_take_snapshot(0, path_to_images, 720, 480)
            jpg_convertion(path_to_images)

#Time stamp function
def convert_format(t):
    mil = t
    seconds=int(mil * 1e-3)
    slider_val = time.strftime("%H:%M:%S", time.gmtime(int(seconds)))
    return f'{slider_val}'

#Player 
def OnTick():
    global player
    global timeSliderLast, timeSliderUpdate
    global factor, org, wro
    # print(factor)
    if player:
        wro = player.get_length() * 1e-3 # to seconds
        factor = wro / org
        # print(t/60)
        end_time = convert_format(int(org*1e3))
        end_slider_label.config(text=end_time)
        if wro > 0:
            timeSlider.config(to=org)
            t = player.get_time() * 1e-3  # to seconds
            
            if t > 0 and t<org and time.time() > (timeSliderUpdate + 2):
                # print(factor)
                # if int(factor) <= 1:
                #     timeSlider.set(t)
                #     timer = convert_format(int(t*1e3))
                #     st_slider_label.config(text= timer)
                # else:
                timeSlider.set(t )
                timer = convert_format(int(t*1e3))
                st_slider_label.config(text= timer)
            timeSliderLast = int(timeVar.get())

    # start the 1 second timer again
    root.after(1000, OnTick)


#selecting and playing vedio:
def start_video():
    global path_to_video
    if path_to_video == "":
        messagebox.showwarning("Warning","Please select the Video file")
    else:
        try:
            global player, timeSliderLast, timeSliderUpdate
            if not player.is_playing():
                timeSliderUpdate = time.time()
                timeSliderLast = 0
                Media = Instance.media_new(path_to_video)
                global org
                clip = VideoFileClip(path_to_video)
                org = clip.duration
                timeSlider.config(to=org)
                player.set_hwnd(vf.winfo_id())
                player.set_media(Media)
                
                em = player.event_manager()
                em.event_attach(vlc.EventType.MediaPlayerEndReached, onEnd)
                player.play()
                _Pause_Play(True)
                OnTick()
            else:
                pass
        except Exception as e:
            print(e)
            messagebox.showwarning("Warning","Error in file name")

def change_speed(*args):
    global player
    speed = speedVar.get()
    player.set_rate(speed)

def onEnd(event):
    global you_cannot_pass, player
    if event.type == vlc.EventType.MediaPlayerEndReached:
        # player.set_time(0)
        print("event")
        you_cannot_pass = True
        # player.stop()

#Toggles play and pause
def _Pause_Play(playing):
    p = 'PAUSE' if playing else 'PLAY'
    global _stopped
    _stopped = False
    pause_btn.config(text=p)

#Pause the video
def pause_video(*args):
    global player
    if player.get_media():
            _Pause_Play(not player.is_playing())
            player.pause()  

#Stop the video
def stop_video():
    global _stopped, path_to_video
    if player:
        player.stop()
        # player.vlm_del_media(path_to_video)
        _Pause_Play(None)
        timeSlider.set(0)
        _stopped = True

# when we change the slider, this function envokes
def OnTime(*unused):
    
    global player
    global timeSliderLast, timeSliderUpdate, you_cannot_pass
    if player:
        global factor, wro, org
        # wro = player.get_length() #ms
        # print(wro, org)
        # factor = wro / org
        t = timeVar.get()
        # print(t)
        if timeSliderLast != int(t):
            player.set_time(int(t * factor* 1e3))  # milliseconds
            timer = convert_format( int(t*1e3))
            # print(player.is_playing())
            st_slider_label.config(text=timer)
            timeSliderUpdate = time.time()
            if you_cannot_pass:
                start_video()
                you_cannot_pass = False

#adding to excel
def add_to_excel():
    global path_to_video
    for i in mylist.curselection():
        code_name = mylist.get(i)
    if code_name == "":
        messagebox.showwarning("Warning","Please select the Defect CODE")
    else :
        if path_to_video == "":
            messagebox.showwarning("Warning","Please select the Video file")
        else:
            pt = path_to_video.split("/")
            video_name = pt[-1].split(".")
            video = video_name[0]
            code = code_name
            try:
                wb_obj = openpyxl.load_workbook("defects_count.xlsx")
                # wb_obj.close()
                # wb_obj = openpyxl.load_workbook("defects_count.xlsx")
            except:
                wb = openpyxl.Workbook()
                wb.save("defects_count.xlsx")
                wb.close()
                wb_obj = openpyxl.load_workbook("defects_count.xlsx")
            wb = wb_obj.active
            c = wb.cell(row = 1, column = 1)
            c.value = "Code"
            defect_codes = []
            video_names = []
            no_of_rows = wb.max_row
            no_of_column = wb.max_column
            for i in range(1,no_of_rows+1):
                defect_codes.append(wb.cell(row=i, column=1).value)
            for i in range(1, no_of_column+1):
                video_names.append(wb.cell(row=1, column=i).value)

            
            if code not in defect_codes: # (checking for code name is alreary there and not present)
                c1 = wb.cell(row=no_of_rows+1, column=1)
                c1.value = code # (code name created)
                if video not in video_names: ## (both code and video not avalilable)
                    c2 = wb.cell(row=1, column=no_of_column+1)
                    c2.value = video #(Video name created)
                    c2.alignment = Alignment(wrapText=True)
                    val = wb.cell(row=no_of_rows+1,column=no_of_column+1).value
                    c3 = wb.cell(row=no_of_rows+1,column=no_of_column+1)
                    if val is None: #(checks wether value is NONE or not)
                        c3.value = 1
                    else:
                        c3.value = val+1
                else: # (but if video name is there)
                    pos = video_names.index(video)
                    val = wb.cell(row=no_of_rows+1,column=pos+1).value
                    c4 = wb.cell(row = no_of_rows+1, column=pos+1)
                    if val is None:
                        c4.value = 1
                    else:
                        c4.value = val+1

            else: #( code name is there ) )
                pos1 = defect_codes.index(code)
                if video not in video_names:#(but no vedio name)
                    c2 = wb.cell(row=1, column=no_of_column+1)
                    c2.value = video
                    c2.alignment = Alignment(wrapText=True)
                    val = wb.cell(row=pos1+1,column=no_of_column+1).value
                    c3 = wb.cell(row=pos1+1,column=no_of_column+1)
                    if val is None:
                        c3.value = 1
                    else:
                        c3.value = val+1
                    c3.alignment = Alignment(wrapText=True)
                
                else: # (both code name nad vedio name are there)
                    pos2 = video_names.index(video)
                    val = wb.cell(row=pos1+1,column=pos2+1).value
                    c4 = wb.cell(row = pos1+1, column=pos2+1)
                    if val is None:
                        c4.value = 1
                    else:
                        c4.value = val+1
            wb_obj.save("defects_count.xlsx")
            wb_obj.close()




df = Frame(root)
df.pack(side = RIGHT, fill=Y)

_var = StringVar()
_var.trace("w", lambda name, index, mode, x=_var: var_changed())


#search
search = Entry (df,bd=5,textvariable= _var) 
search.pack(side=TOP, fill  = X)


#Scoller for list box

sb = Scrollbar(df)
sb.pack(side = RIGHT, fill=Y)

global factor
factor = 1

# GUI - defects code list box
mylist = Listbox(df, yscrollcommand = sb.set ) 

for line in defects:  
    mylist.insert(END, line)  
mylist.pack( side = RIGHT,fill=Y )  
sb.config( command = mylist.yview )  

# video frame
vf=Frame(root,  bg = 'red',height=500, width=1000)
vf.pack(fill=None, expand=False, pady=10)

# Button for contols
cf = Frame(root)
cf.pack()
snap_btn = Button(cf, pady=10,padx=5, text='SNAP', command=selected_item)
snap_btn.grid(row = 0, column=0, padx=10)
play_btn = Button(cf, pady=10,padx=5, text='START', command=start_video)
play_btn.grid(row = 0, column=1, padx=10)
pause_btn = Button(cf, pady=10,padx=5, text='PAUSE', command = pause_video)
pause_btn.grid(row = 0, column=2, padx=10)
stop_btn = Button(cf, pady=10,padx=5, text='STOP', command=stop_video)
stop_btn.grid(row = 0, column=3, padx=10)
# speed_btn = Button(cf, pady=10,padx=10, text='1x', command=change_speed)
# speed_btn.grid(row = 0, column=4, padx=10)
stop_btn = Button(cf, pady=10,padx=5, text='STOP', command=stop_video)
stop_btn.grid(row = 0, column=3, padx=10)
speed_label = Label(cf, pady=10,padx=5, text='SPEED')
speed_label.grid(row = 0, column=4, padx=1)
speedVar = DoubleVar()
speedVar.set(1)
speedSlider = Scale(cf, variable=speedVar, command=change_speed,
                                from_=0.25, to=3, orient=HORIZONTAL, length=100,
                                digits     =   3,         # MVC-Visual-Part presentation trick
                                resolution =   0.25 ,
                                )  # label='Time',
speedSlider.grid(row = 0, column=5, padx=10)
excel_btn = Button(cf, pady=10,padx=5, text='ADD TO EXCEL', command=add_to_excel)
excel_btn.grid(row=0, column=6 ,padx=100 )

#Slider timer
timers = Frame(root)
timeVar = DoubleVar()
global timeSliderLast
timeSliderLast = 0
timeSlider = Scale(timers, variable=timeVar, command=OnTime,
                                from_=0, to=1000, orient=HORIZONTAL, length=500,
                                showvalue=0)  # label='Time',
timeSlider.pack(side=BOTTOM, fill=X, expand=1)
global timeSliderUpdate
timeSliderUpdate = time.time()
timers.pack(side=BOTTOM, fill=X, pady=30, padx=25)

# Slider Label
st_slider_label = Label(timers,text = "00:00:00")
st_slider_label.pack(side= LEFT,pady=10)
end_slider_label = Label(timers,text = "00:00:00")
end_slider_label.pack(side= RIGHT,pady=10)

#creating menu
my_menu = Menu(root)
root.config(menu=my_menu)
#Defects menu
folder_path_menu= Menu(my_menu)
my_menu.add_cascade(label="Defects",menu = folder_path_menu)
folder_path_menu.add_command(label="Defect Folder Path", command=add_path_folder)
#Video menu
video_path_menu= Menu(my_menu)
my_menu.add_cascade(label="Video",menu = video_path_menu)
video_path_menu.add_command(label="Video Path", command=add_video_file)


df.bind("<1>", lambda event: search.focus_set())
vf.bind("<1>", lambda event: vf.focus_set())
cf.bind("<1>", lambda event: cf.focus_set())
# root.bind('<1>', lambda event: root.focus_set())
root.bind('<space>',pause_video)
root.bind('S', selected_item)
root.mainloop()
